from django.core.exceptions import ValidationError
from datetime import datetime
from django.shortcuts import render
import base64
from django.core.files.base import ContentFile
from core.models import (
    order_transactions, operator_code_master,
    admin_code_master, product_info_master,
    product_category_master, OrderTransactionPDF, OrderTransactionItemPDF,
    OrderTransactionExcel, OrderTransactionItemExcel
)
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login as login_function
from django.template.loader import render_to_string

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, time
import json
import pytz
from django.core.exceptions import ValidationError
from weasyprint import HTML, CSS

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
# Flag to track if settlement date has been processed for the day
# settlement_processed_today = False
# Create your views here.


def front(request):
    context = {}
    return render(request, "index.html", context)


def get_order_transactions(request, data_flag):
    if request.method == "GET":
        if data_flag == "all":
            data = order_transactions.objects.filter(
                product_code=1).filter(status="active").values()
            print(data)
        else:
            data = order_transactions.objects.filter(product_code=1).filter(provision_completion_flag=False).filter(
                order_cancellation_flag=False).filter(status="active").values()
        return JsonResponse(list(data), safe=False)


@csrf_exempt
def change_status(request, order_id):
    if request.method == "GET":
        row = order_transactions.objects.filter(
            order_no=order_id).update(provision_completion_flag=True, provider_operator_code=request.session.get('operator_code', ''))
        return HttpResponse("Order has been updated")


@csrf_exempt
def cancel_order(request, order_id):
    if request.method == "GET":
        row = order_transactions.objects.filter(
            order_no=order_id).update(order_cancellation_flag=True)
        return HttpResponse("Order has been updated")


@csrf_exempt
def login(request):
    if request.method == "POST":
        body = json.loads(request.body)
        operator_code = body['operator_code']
        operator_password = body['operator_password']
        check_user = operator_code_master.objects.filter(
            operator_password=operator_password).filter(
            operator_code=operator_code).first()
        if check_user:
            request.session['operator_code'] = check_user.operator_code
            return HttpResponse(True)
        else:
            return HttpResponse(False)


@csrf_exempt
def logout(request):
    if request.method == "GET":
        return HttpResponse("Order has been updated")


@csrf_exempt
def register(request):
    if request.method == "POST":
        body = json.loads(request.body)
        operator_name = body['operator_name']
        password = body['password']
        operator_code = body['operator_code']
        new_user = operator_code_master(
            operator_code=operator_code, operator_name=operator_name, operator_password=password, last_login="")
        new_user.save()
        return HttpResponse(True)


@csrf_exempt
def adminlogin(request):
    if request.method == "POST":
        body = json.loads(request.body)
        admin_code = body['admin_code']
        admin_password = body['admin_password']
        check_user = admin_code_master.objects.filter(
            admin_password=admin_password).filter(
            admin_code=admin_code).first()
        if check_user:
            request.session['admin_code'] = check_user.admin_code
            return HttpResponse(True)
        else:
            return HttpResponse(False)


@csrf_exempt
def adminlogout(request):
    if request.method == "GET":
        return HttpResponse("Order has been updated")


@csrf_exempt
def adminregister(request):
    if request.method == "POST":
        body = json.loads(request.body)
        admin_name = body['admin_name']
        password = body['password']
        admin_code = body['admin_code']
        new_user = admin_code_master(
            admin_code=admin_code, admin_name=admin_name, admin_password=password, last_login="")
        new_user.save()
        return HttpResponse(True)


@csrf_exempt
def add_product(request):
    if request.method == 'POST':
        product_id = request.POST['product_id']
        product_name = request.POST['product_name']
        product_price = request.POST['product_price']
        product_category = request.POST['product_category']
        image = request.POST['image']
        new_product = product_info_master(product_id=product_id, product_name_en=product_name,
                                          product_price_en=product_price, category_name=product_category, product_image_link_dest=image)
        new_product.save()
        return HttpResponse(True)


@csrf_exempt
def product_data(request):
    if request.method == 'GET':
        product_data = product_info_master.objects.values()
        return JsonResponse(list(product_data), safe=False)


@csrf_exempt
def get_admin(request):
    if request.method == 'GET':
        return HttpResponse(request.session.get('admin_code', ''))


@csrf_exempt
def get_category(request):
    if request.method == 'GET':
        return JsonResponse([i for i in product_category_master.objects.values_list('product_category')], safe=False)


@csrf_exempt
def add_category(request):
    if request.method == 'POST':
        product_category_no = request.POST['product_category_no']
        product_category = request.POST['product_category']
        product_category_name_en = request.POST['product_category_name_en']
        new_category = product_category_master(product_category_no=product_category_no, product_category=product_category,
                                               product_category_name_en=product_category_name_en)
        new_category.save()
        return HttpResponse(True)


@csrf_exempt
def get_product_category(request):
    if request.method == 'GET':
        product_data = product_category_master.objects.values()
        return JsonResponse(list(product_data), safe=False)


@csrf_exempt
def generate_pdf(request):
    order_transactions_list = []
    total_order = 0
    total_amount = 0

    # Fetch all orders with settlement dates
    orders = order_transactions.objects.filter(settlement_date__isnull=False,
                                               provision_completion_flag=False,
                                               order_cancellation_flag=False,
                                               status="active")

    for order in orders:
        # Extract the order_time from the order
        order_time = order.order_time.date()

        # Compare selected_date and order_time
        if order.settlement_date == order_time:
            order_transactions_list.append([
                order.lane_no,
                order.table_no,
                order.order_amount,
                order.product_unit_price * order.order_amount,
                order.settlement_date.strftime('%Y-%m-%d'),
            ])

            total_order += order.order_amount
            total_amount += order.product_unit_price * order.order_amount

    # Create PDF document
    pdf_filename = f"Settlements_Report_Table_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    pdf_file_path = f"media/{pdf_filename}"
    doc = SimpleDocTemplate(pdf_file_path, pagesize=letter)

    # Define table data
    table_data = [["LAN No.", "Table No.", "Total Order",
                   "Total Amount", "Settlement Date"]]
    table_data.extend(order_transactions_list)
    table_data.append(["", "", "Total Order", total_order, ""])
    table_data.append(["", "", "Total Amount", total_amount, ""])

    # Create table
    table = Table(table_data, colWidths=[70, 70, 70, 70, 100])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

    # Add table to the PDF document
    doc.build([table])

    # Save file path to database
    order_transaction_obj = OrderTransactionPDF.objects.create(
        transaction_time=datetime.now(), pdf_file=pdf_filename)

    return JsonResponse({'pdf_url': order_transaction_obj.pdf_file.url}, safe=False)


@csrf_exempt
def generate_pdf_item(request):
    order_transactions_list = []
    total_order = 0
    total_amount = 0

    # Fetch all orders with settlement dates
    orders = order_transactions.objects.filter(settlement_date__isnull=False,
                                               provision_completion_flag=False,
                                               order_cancellation_flag=False,
                                               status="active")

    for order in orders:
        # Extract the order_time from the order
        order_time = order.order_time.date()

        # Compare selected_date and order_time
        if order.settlement_date == order_time:
            order_transactions_list.append([
                order.product_code,
                order.product_name_en,
                order.order_amount,
                order.product_unit_price * order.order_amount,
                order.settlement_date.strftime('%Y-%m-%d'),
            ])

            total_order += order.order_amount
            total_amount += order.product_unit_price * order.order_amount

    # Create PDF document
    pdf_filename = f"Settlements_Report_Product_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    pdf_file_path = f"media/{pdf_filename}"
    doc = SimpleDocTemplate(pdf_file_path, pagesize=letter)

    # Define table data
    table_data = [["Product Code", "Product Name.",
                   "Total Order", "Total Amount", "Settlement Date"]]
    table_data.extend(order_transactions_list)
    table_data.append(["", "", "Total Order", total_order, ""])
    table_data.append(["", "", "Total Amount", total_amount, ""])

    # Create table
    table = Table(table_data, colWidths=[70, 150, 70, 70, 100])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

    # Add table to the PDF document
    doc.build([table])

    # Save file path to database
    order_transaction_obj = OrderTransactionItemPDF.objects.create(
        transaction_time=datetime.now(), pdf_file=pdf_filename)

    return JsonResponse({'pdf_url': order_transaction_obj.pdf_file.url}, safe=False)


@csrf_exempt
def search_pdf(request):
    if request.method == "GET":
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        pdf_url_list = []
        from_date = datetime.strptime(from_date, '%Y-%m-%d')
        to_date = datetime.strptime(to_date, '%Y-%m-%d')
        from_date = datetime.combine(from_date.date(),
                                     time(0, 0, 0, tzinfo=pytz.timezone('UTC')))
        to_date = datetime.combine(to_date.date(),
                                   time(23, 59, 59, tzinfo=pytz.timezone('UTC')))
        order_transaction_pdfs = OrderTransactionPDF.objects.filter(
            transaction_time__range=[from_date, to_date]
        ).order_by('-transaction_time')
        for pdf in order_transaction_pdfs:
            pdf_url_list.append({
                'pdf_url': pdf.pdf_file.url,
                'transaction_time': pdf.transaction_time
            })
        return JsonResponse({'pdf_url_list': pdf_url_list}, safe=False)


@csrf_exempt
def search_item_pdf(request):
    if request.method == "GET":
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        pdf_url_list = []
        from_date = datetime.strptime(from_date, '%Y-%m-%d')
        to_date = datetime.strptime(to_date, '%Y-%m-%d')
        from_date = datetime.combine(from_date.date(),
                                     time(0, 0, 0, tzinfo=pytz.timezone('UTC')))
        to_date = datetime.combine(to_date.date(),
                                   time(23, 59, 59, tzinfo=pytz.timezone('UTC')))
        order_transaction_pdfs = OrderTransactionItemPDF.objects.filter(
            transaction_time__range=[from_date, to_date]
        ).order_by('-transaction_time')
        for pdf in order_transaction_pdfs:
            pdf_url_list.append({
                'pdf_url': pdf.pdf_file.url,
                'transaction_time': pdf.transaction_time
            })
        return JsonResponse({'pdf_url_list': pdf_url_list}, safe=False)


@csrf_exempt
def generate_excel(request):
    order_transactions_list = []
    total_order = 0
    total_amount = 0

    # Fetch all orders with settlement dates
    orders = order_transactions.objects.filter(settlement_date__isnull=False,
                                               provision_completion_flag=False,
                                               order_cancellation_flag=False,
                                               status="active")

    for order in orders:
        # Extract the order_time from the order
        order_time = order.order_time.date()

        # Compare selected_date and order_time
        if order.settlement_date == order_time:
            order_transactions_list.append({
                'lane_no': order.lane_no,
                'table_no': order.table_no,
                'total_order': order.order_amount,
                'total_amount': order.product_unit_price * order.order_amount,
                'settlement_date': order.settlement_date.strftime('%Y-%m-%d'),
            })

            total_order += order.order_amount
            total_amount += order.product_unit_price * order.order_amount

    # Generate Excel file
    wb = Workbook()
    ws = wb.active

    # Add headers
    headers = ["LAN No.", "Table No.", "Total Order",
               "Total Amount", "Settlement Date"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

    # Add data
    for row_idx, transaction in enumerate(order_transactions_list, start=2):
        ws.cell(row=row_idx, column=1, value=transaction["lane_no"])
        ws.cell(row=row_idx, column=2, value=transaction["table_no"])
        ws.cell(row=row_idx, column=3, value=transaction["total_order"])
        ws.cell(row=row_idx, column=4, value=transaction["total_amount"])
        ws.cell(row=row_idx, column=5, value=transaction["settlement_date"])

    # Add total order and total amount rows
    total_order_row = len(order_transactions_list) + 2
    ws.cell(row=total_order_row, column=2, value="Total Order")
    ws.cell(row=total_order_row, column=3, value=total_order)

    total_amount_row = len(order_transactions_list) + 3
    ws.cell(row=total_amount_row, column=2, value="Total Amount")
    ws.cell(row=total_amount_row, column=3, value=total_amount)

    # Save the Excel file
    filename = f"Settlements_Report_Table_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = f"media/{filename}"
    wb.save(file_path)

    # Save file path to database
    order_transaction_obj = OrderTransactionExcel.objects.create(
        transaction_time=datetime.now(), excel_file=filename)

    return JsonResponse({'excel_url': order_transaction_obj.excel_file.url}, safe=False)


@csrf_exempt
def generate_Item_excel(request):
    order_transactions_list = []
    total_order = 0
    total_amount = 0

    # Fetch all orders with settlement dates
    orders = order_transactions.objects.filter(settlement_date__isnull=False,
                                               provision_completion_flag=False,
                                               order_cancellation_flag=False,
                                               status="active")

    for order in orders:
        # Extract the order_time from the order
        order_time = order.order_time.date()

        # Compare selected_date and order_time
        if order.settlement_date == order_time:
            order_transactions_list.append({
                'product_code': order.product_code,
                'product_name_en': order.product_name_en,
                'total_order': order.order_amount,
                'total_amount': order.product_unit_price * order.order_amount,
                'settlement_date': order.settlement_date.strftime('%Y-%m-%d'),
            })

            total_order += order.order_amount
            total_amount += order.product_unit_price * order.order_amount

    # Generate Excel file
    wb = Workbook()
    ws = wb.active

    # Add headers
    headers = ["Product Code", "Product Name.", "Total Order",
               "Total Amount", "Settlement Date"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

    # Add data
    for row_idx, transaction in enumerate(order_transactions_list, start=2):
        ws.cell(row=row_idx, column=1, value=transaction["product_code"])
        ws.cell(row=row_idx, column=2, value=transaction["product_name_en"])
        ws.cell(row=row_idx, column=3, value=transaction["total_order"])
        ws.cell(row=row_idx, column=4, value=transaction["total_amount"])
        ws.cell(row=row_idx, column=5, value=transaction["settlement_date"])

    # Add total order and total amount rows
    total_order_row = len(order_transactions_list) + 2
    ws.cell(row=total_order_row, column=2, value="Total Order")
    ws.cell(row=total_order_row, column=3, value=total_order)

    total_amount_row = len(order_transactions_list) + 3
    ws.cell(row=total_amount_row, column=2, value="Total Amount")
    ws.cell(row=total_amount_row, column=3, value=total_amount)

    # Save the Excel file
    filename = f"Settlements_Report_Product_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = f"media/{filename}"
    wb.save(file_path)

    # Save file path to database
    order_transaction_obj = OrderTransactionItemExcel.objects.create(
        transaction_time=datetime.now(), excel_file=filename)

    return JsonResponse({'excel_url': order_transaction_obj.excel_file.url}, safe=False)


@csrf_exempt
def search_excel(request):
    if request.method == "GET":
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        excel_url_list = []
        from_date = datetime.strptime(from_date, '%Y-%m-%d')
        to_date = datetime.strptime(to_date, '%Y-%m-%d')
        from_date = datetime.combine(from_date.date(),
                                     time(0, 0, 0, tzinfo=pytz.timezone('UTC')))
        to_date = datetime.combine(to_date.date(),
                                   time(23, 59, 59, tzinfo=pytz.timezone('UTC')))
        order_transaction_excels = OrderTransactionExcel.objects.filter(
            transaction_time__range=[from_date, to_date]
        ).order_by('-transaction_time')
        for excel in order_transaction_excels:
            excel_url_list.append({
                'excel_url': excel.excel_file.url,
                'transaction_time': excel.transaction_time
            })
        return JsonResponse({'excel_url_list': excel_url_list}, safe=False)


@csrf_exempt
def search_Item_excel(request):
    if request.method == "GET":
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        excel_url_list = []
        from_date = datetime.strptime(from_date, '%Y-%m-%d')
        to_date = datetime.strptime(to_date, '%Y-%m-%d')
        from_date = datetime.combine(from_date.date(),
                                     time(0, 0, 0, tzinfo=pytz.timezone('UTC')))
        to_date = datetime.combine(to_date.date(),
                                   time(23, 59, 59, tzinfo=pytz.timezone('UTC')))
        order_transaction_excels = OrderTransactionItemExcel.objects.filter(
            transaction_time__range=[from_date, to_date]
        ).order_by('-transaction_time')
        for excel in order_transaction_excels:
            excel_url_list.append({
                'excel_url': excel.excel_file.url,
                'transaction_time': excel.transaction_time
            })
        return JsonResponse({'excel_url_list': excel_url_list}, safe=False)


@csrf_exempt
def get_order_date(request):
    if request.method == 'GET':
        order_time = order_transactions.objects.filter(order_no=1).filter(provision_completion_flag=False).filter(
            order_cancellation_flag=False).filter(status="active").values_list('order_time')
        return HttpResponse(order_time[0][0].strftime("%a %b %d %Y %H:%M:%S GMT%z (%Z)"))


# @csrf_exempt
# def store_settlement_date(request):
#     if request.method == 'POST':
#         body = json.loads(request.body)
#         settlement_date = body['settlement_date']
#         order_transactions.objects.filter(order_no=1).filter(provision_completion_flag=False).filter(
#             order_cancellation_flag=False).filter(status="active").update(settlement_date=datetime.strptime(settlement_date, '%Y-%m-%dT%H:%M:%S.%fZ'))
#         return JsonResponse(True, safe=False)


@csrf_exempt
def store_settlement_date(request):
    if request.method == 'POST':
        body = json.loads(request.body)
        settlement_date_str = body.get('settlement_date')

        try:
            # Attempt to parse the settlement_date string
            settlement_date = None

            # Try parsing the date with different formats
            for date_format in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S.%fZ'):
                try:
                    settlement_date = datetime.strptime(
                        settlement_date_str, date_format).date()
                    break  # Break the loop if parsing succeeds
                except ValueError:
                    pass  # Continue trying other formats

            if settlement_date is None:
                raise ValueError('Invalid date format')

        except ValueError as e:
            # Handle the case where the date format is invalid
            return JsonResponse({'error': str(e)}, status=400)

        # Get all orders
        orders = order_transactions.objects.filter(provision_completion_flag=False,
                                                   order_cancellation_flag=False,
                                                   status="active")

        if not orders:
            return JsonResponse({'error': 'No active orders found'}, status=404)

        # Iterate over each order
        settlement_updated = False
        for order in orders:
            # Extract the order_time from the order
            order_time = order.order_time.date()

            # Compare selected_date and order_time
            selected_date = settlement_date
            if selected_date == order_time:
                # Update settlement_date for the order
                order.settlement_date = settlement_date
                order.save()
                settlement_updated = True

        if settlement_updated:
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': 'No matching orders found'}, status=404)


# @csrf_exempt
# def store_settlement_date(request):
#     if request.method == 'POST':
#         body = json.loads(request.body)
#         settlement_date = body.get('settlement_date')
#         if not settlement_date:
#             return JsonResponse({'error': 'Settlement date not provided'}, status=400)

#         # Fetch orders based on the order_time value
#         order_time = body.get('order_time')
#         if not order_time:
#             return JsonResponse({'error': 'Order time not provided'}, status=400)

#         # Fetch orders with the same order_time
#         orders = order_transactions.objects.filter(order_time=order_time, provision_completion_flag=False,
#                                                    order_cancellation_flag=False, status="active")

#         # Check if any orders are found
#         if not orders:
#             return JsonResponse({'error': 'Orders not found'}, status=404)

#         # Update the settlement date for each matching order in the database
#         for order in orders:
#             order.settlement_date = settlement_date
#             order.save()

#         return JsonResponse({'success': True})
